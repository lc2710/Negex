import re
from openpyxl import load_workbook
#import csv
#import numpy as np

def format_file():

    keywords = []
    keywordVals = []

    #Reading in Keywords
    #wb = load_workbook(filename='Input Files/Grooming_Categories.xlsx', read_only=True)
    wb = load_workbook(filename='Service Provider Categorise/Grooming_Categories.xlsx', read_only=True)
    ws = wb['Sheet1'] # ws is now an IterableWorksheet

    MFList = []
    MFArray = []
    category_count = 0
    i = 1

    for row in ws.iter_rows():

        i = i + 1 ##Count columns on second row iteration to avoid bugs in empty columns

        for cell in row:

            if i == 2:
                category_count = category_count + 1

    category_count = category_count - 1

    for row in ws.rows:

        keywords.append(str(row[0].value).upper())

    #print(keywordVals)
    category = 0    #Number for each descriptive category

    #Reading in review file
    myfile = open(r'Input Files/zappatas_review.txt', 'r')
    myfile = myfile.readlines()
    myfile = str(myfile)

    myfile = re.sub(r'\[', '',myfile)
    myfile = re.sub(r'\',', '',myfile)
    myfile = re.sub(r'\'', '',myfile)
    myfile = re.sub(r'\",', '',myfile)
    myfile = re.sub(r'\"]', '',myfile)
    myfile = re.sub(r'\\n', '',myfile)
    myfile = re.sub(r'\"', '',myfile)

    myfile = str.split(str(myfile), '.')
    #print(myfile)

    #Formatting input file for Negex algorithm
    f = open('negex_in.txt','w')
    f.write('Report No.\tConcept\tSentence\tNegation\n')
    text_out=""
    wordmatch = 0
    no_keyword_list = []
    no_keyword_text = ""

    for line in myfile:

        #Find keyword matches
        keyword_match = []
        for word in str.split(line):

            if word.upper() in keywords:

                keyword_match.append(word.upper())
                wordmatch = wordmatch + 1
                #print(word.upper())
                MFArray = []

                #Creating MF value matrix
                #for col in range(1, category_count-1):
                for col in range(2, category_count-1):      #Not including keywords, only MF values

                    #print(ws.cell(row=keywords.index(word.upper())+1, column=col).value)
                    MFArray = MFArray + [ws.cell(row=keywords.index(word.upper())+1, column=col).value]

                MFList.append(MFArray)

        if wordmatch > 0:
            #print(wordmatch)
            pass

            for word in keyword_match:
                text_out += str(category) + '\t'        #Category = number representing each category - need to fill this out
                text_out += word + '\t'
                text_out += line + '.\t' + 'Dummytext' + '\n'

        else:
            #Only need one of these
            no_keyword_list.append(line)
            no_keyword_text += line + '\n'

        wordmatch = 0
        keyword_match = []

    no_keyword_text = re.sub(r'\[', '',no_keyword_text)
    no_keyword_text = re.sub(r'\',', '',no_keyword_text)
    no_keyword_text = re.sub(r'\'', '',no_keyword_text)
    no_keyword_text = re.sub(r'\",', '',no_keyword_text)
    no_keyword_text = re.sub(r'\"]', '',no_keyword_text)
    no_keyword_text = re.sub(r'\\n', '',no_keyword_text)

    f.write(text_out)
    f.close()
    #print(text_out)
    #print(no_keyword_text)         #Scan through this data to find new keywords??

    print("MFList: " + str(MFList))

    descriptCategories = []
    for col in range(2, category_count-1):      #Not including keywords, only MF values

        descriptCategories = descriptCategories + [ws.cell(row=1, column=col).value]

    print(descriptCategories)
    return MFList, descriptCategories

#format_file()